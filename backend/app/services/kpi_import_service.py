"""
Импорт KPI_Mapping.xlsx → таблицы kpi_indicators, kpi_criteria, kpi_role_cards,
kpi_role_card_indicators.

Логика дедупликации:
- is_common показатели: один indicator+criterion на весь справочник
  (ключ дедупликации = criterion.strip())
- Не-is_common: один indicator на уникальную пару (name, formula_type);
  критерий создаётся на каждую уникальную тройку (indicator_id, criterion, weight)
- Карточки: одна per pos_id (active статус при импорте)
"""

import logging
import os
import re
import uuid
from datetime import date, datetime, timezone
from typing import Optional

import openpyxl

from app.services.report_service import get_common_kpi_texts
from app.services.threshold_parser import parse_thresholds

logger = logging.getLogger(__name__)

KPI_MAPPING_PATH = os.environ.get("KPI_MAPPING_PATH", "/app/reference/KPI_Mapping.xlsx")


def _norm(text: str) -> str:
    """Нормализация текста для ключей дедупликации: убирает лишние пробелы и переносы строк."""
    return re.sub(r'\s+', ' ', text.strip())

INDICATOR_GROUPS = {
    "Общие показатели": None,
    "Проектная деятельность": ["проект", "трансформац", "цифров"],
    "Аналитическая деятельность": ["аналитич", "анализ", "мониторинг", "отчёт", "отчет"],
    "Закупочная деятельность": ["закуп", "торг", "контракт", "поставщ"],
    "Правовое обеспечение": ["правов", "юридич", "закон", "норматив"],
    "Документооборот": ["документ", "обращени", "МСЭД", "ЗК МСЭД"],
    "Информационные технологии": ["информацион", "техническ", "систем", "IT", "ИТ", "ЕАСУЗ"],
    "Организационное обеспечение": ["организацион", "бюджет", "финанс", "учёт"],
    "Прочие показатели": [],
}


def _classify_indicator(name: str, is_common: bool) -> str:
    if is_common:
        return "Общие показатели"
    name_lower = name.lower()
    for group, keywords in INDICATOR_GROUPS.items():
        if keywords is None:
            continue
        if not keywords:
            continue
        for kw in keywords:
            if kw.lower() in name_lower:
                return group
    return "Прочие показатели"

# Маппинг is_common критерий → стандартные тексты
_COMMON_FRAGMENTS = {
    "исполнительской дисциплин": None,
    "трудового распорядка": None,
    "техники безопасности": None,
    "охран": None,
    "МСЭД": None,
}


def _is_common_criterion(criterion: str) -> bool:
    cl = criterion.lower()
    for frag in _COMMON_FRAGMENTS:
        if frag.lower() in cl:
            return True
    return False


def _parse_formula_desc(formula_desc: str) -> tuple[Optional[str], Optional[str]]:
    """Пытается извлечь числитель и знаменатель из formula_desc."""
    if not formula_desc:
        return None, None
    # Простой паттерн: "Числитель / Знаменатель" или "Числитель/Знаменатель"
    if "/" in formula_desc:
        parts = formula_desc.split("/", 1)
        return parts[0].strip() or None, parts[1].strip() or None
    return formula_desc.strip() or None, None


def _thresholds_to_json(thresholds_str: str) -> Optional[list[dict]]:
    """Конвертирует строку порогов в JSON-список через ThresholdParser."""
    if not thresholds_str or not thresholds_str.strip():
        return None
    try:
        rules = parse_thresholds(thresholds_str)
        if not rules:
            return None
        return [r.model_dump() for r in rules]
    except Exception:
        return None


class KpiImportService:

    def import_from_xlsx(self, xlsx_path: str = KPI_MAPPING_PATH) -> dict:
        """
        Читает KPI_Mapping.xlsx и возвращает структуры для вставки в БД.
        НЕ делает DB-операций — возвращает dict с данными для вставки.
        Db-операции выполняет async endpoint.
        """
        errors: list[str] = []

        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        except FileNotFoundError:
            return {"error": f"Файл не найден: {xlsx_path}"}
        except Exception as e:
            return {"error": str(e)}

        # Читаем роли
        roles: dict[str, dict] = {}  # role_id → {pos_id, role_name, ...}
        for row in wb["KPI_Roles"].iter_rows(min_row=2, values_only=True):
            if not row[1]:
                continue
            role_id = str(row[1])
            roles[role_id] = {
                "pos_id": int(row[0]) if row[0] else 0,
                "role_id": role_id,
                "role_name": str(row[4]) if row[4] else "",
                "unit": str(row[3]) if row[3] else "",
            }

        # Читаем индикаторы
        raw_rows: list[dict] = []
        for row in wb["KPI_Indicators"].iter_rows(min_row=2, values_only=True):
            if not row[1]:
                continue

            formula_type = str(row[7]).strip() if row[7] else "binary_manual"
            if formula_type not in (
                "binary_auto", "binary_manual", "threshold",
                "multi_threshold", "quarterly_threshold"
            ):
                formula_type = "binary_manual"

            raw_is_common = row[6]
            is_common = raw_is_common if isinstance(raw_is_common, bool) else str(raw_is_common).upper() == "TRUE"

            raw_cumul = row[10]
            cumulative = raw_cumul if isinstance(raw_cumul, bool) else str(raw_cumul).upper() == "TRUE"

            try:
                weight = int(row[5]) if row[5] is not None else 0
            except (ValueError, TypeError):
                weight = 0

            raw_rows.append({
                "pos_id": int(row[0]) if row[0] else 0,
                "role_id": str(row[1]),
                "indicator": str(row[2]) if row[2] else "",
                "criterion": str(row[3]) if row[3] else "",
                "plan_value": str(row[4]) if row[4] else "",
                "weight": weight,
                "is_common": is_common,
                "formula_type": formula_type,
                "formula_desc": str(row[8]) if row[8] else "",
                "thresholds": str(row[9]) if row[9] else "",
                "cumulative": cumulative,
            })

        wb.close()

        today = date.today()

        # ── Дедупликация indicators ───────────────────────────────────────────
        # Ключ для is_common: criterion.strip() (одинаковый для всех ролей)
        # Ключ для обычных: (indicator_name, formula_type)
        indicators: dict[str, dict] = {}          # key → indicator dict
        criteria: dict[str, dict] = {}            # key → criterion dict
        # criterion_key → criterion_id (uuid str)
        criterion_key_to_id: dict[str, str] = {}
        indicator_key_to_id: dict[str, str] = {}

        for r in raw_rows:
            ind_name = r["indicator"]
            formula_type = r["formula_type"]
            criterion_text = r["criterion"]
            is_common = r["is_common"]

            # --- Indicator ---
            if is_common:
                ind_key = f"common::{_norm(criterion_text)}"
                ind_name_stored = _norm(criterion_text)
            else:
                ind_key = f"{_norm(ind_name)}::{formula_type}"
                ind_name_stored = ind_name

            if ind_key not in indicators:
                ind_id = str(uuid.uuid4())
                indicators[ind_key] = {
                    "id": ind_id,
                    "name": ind_name_stored,
                    "formula_type": formula_type,
                    "is_common": is_common,
                    "is_editable_per_role": not is_common,
                    "status": "active",
                    "version": 1,
                    "valid_from": today,
                    "valid_to": None,
                    "created_by": "import",
                    "created_at": datetime.now(timezone.utc),
                    "updated_at": datetime.now(timezone.utc),
                    "indicator_group": _classify_indicator(ind_name_stored, is_common),
                }
                indicator_key_to_id[ind_key] = ind_id

            ind_id = indicator_key_to_id[ind_key]

            # --- Criterion ---
            # Ключ: (indicator_id, criterion_text)
            crit_key = f"{ind_id}::{_norm(criterion_text)}"
            if crit_key not in criteria:
                crit_id = str(uuid.uuid4())
                num_label, den_label = _parse_formula_desc(r["formula_desc"])
                thresholds_json = _thresholds_to_json(r["thresholds"])

                pos_text, neg_text = None, None
                if is_common:
                    pos_text, neg_text = get_common_kpi_texts(criterion_text)

                criteria[crit_key] = {
                    "id": crit_id,
                    "indicator_id": ind_id,
                    "criterion": criterion_text,
                    "numerator_label": num_label,
                    "denominator_label": den_label,
                    "thresholds": thresholds_json,
                    "sub_indicators": None,
                    "quarterly_thresholds": None,
                    "cumulative": r["cumulative"],
                    "plan_value": r["plan_value"] or None,
                    "common_text_positive": pos_text,
                    "common_text_negative": neg_text,
                    "created_at": datetime.now(timezone.utc),
                }
                criterion_key_to_id[crit_key] = crit_id

            # Сохраняем ссылки для построения card_indicators
            r["_ind_id"] = ind_id
            r["_ind_key"] = ind_key
            r["_crit_id"] = criterion_key_to_id[crit_key]

        # ── Карточки должностей ───────────────────────────────────────────────
        # Группируем raw_rows по role_id
        cards: dict[str, dict] = {}              # role_id → card dict
        card_indicators: list[dict] = []

        for r in raw_rows:
            role_id = r["role_id"]
            role_info = roles.get(role_id, {})
            pos_id = role_info.get("pos_id", r["pos_id"])

            if role_id not in cards:
                card_id = str(uuid.uuid4())
                cards[role_id] = {
                    "id": card_id,
                    "pos_id": pos_id,
                    "role_id": role_id,
                    "role_name": role_info.get("role_name", ""),
                    "unit": role_info.get("unit", ""),
                    "version": 1,
                    "status": "active",
                    "valid_from": today,
                    "valid_to": None,
                    "created_by": "import",
                    "approved_by": None,
                    "approved_at": None,
                    "created_at": datetime.now(timezone.utc),
                    "updated_at": datetime.now(timezone.utc),
                    "_used_indicators": set(),  # для дедупликации
                    "_order": 0,
                }

            card = cards[role_id]
            ind_id = r["_ind_id"]

            # Дедупликация: один indicator per card
            if ind_id in card["_used_indicators"]:
                continue
            card["_used_indicators"].add(ind_id)

            card["_order"] += 1
            card_indicators.append({
                "id": str(uuid.uuid4()),
                "card_id": card["id"],
                "indicator_id": ind_id,
                "criterion_id": r["_crit_id"],
                "weight": r["weight"],
                "order_num": card["_order"],
                "override_criterion": None,
                "override_thresholds": None,
                "override_weight": None,
            })

        # Очищаем служебные поля
        for c in cards.values():
            c.pop("_used_indicators", None)
            c.pop("_order", None)

        return {
            "indicators": list(indicators.values()),
            "criteria": list(criteria.values()),
            "cards": list(cards.values()),
            "card_indicators": card_indicators,
            "errors": errors,
        }


kpi_import_service = KpiImportService()
