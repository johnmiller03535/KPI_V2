import logging
import os
from typing import Optional
from functools import lru_cache
import openpyxl

from app.schemas.kpi import KpiItem, KpiStructure

logger = logging.getLogger(__name__)

KPI_MAPPING_PATH = os.environ.get("KPI_MAPPING_PATH", "/app/reference/KPI_Mapping.xlsx")

# formula_type → kpi_type
_FORMULA_TO_KPI_TYPE = {
    "binary_auto":         "binary_auto",
    "binary_manual":       "binary_manual",
    "threshold":           "numeric",
    "multi_threshold":     "numeric",
    "quarterly_threshold": "numeric",
    "absolute_threshold":  "numeric",
}


class KpiMappingService:
    """
    Читает KPI_Mapping.xlsx и предоставляет структуру KPI по должности.

    Лист KPI_Roles (col): pos_id | role_id | management | unit | role
    Лист KPI_Indicators (col): pos_id | role_id | indicator | criterion |
        plan_value | weight | is_common | formula_type | formula_desc |
        thresholds | cumulative
    """

    def __init__(self):
        self._roles: dict[str, dict] = {}
        self._indicators: dict[str, list[KpiItem]] = {}
        self._loaded = False

    def _load(self):
        if self._loaded:
            return
        try:
            wb = openpyxl.load_workbook(KPI_MAPPING_PATH, read_only=True, data_only=True)

            # --- KPI_Roles ---
            for row in wb["KPI_Roles"].iter_rows(min_row=2, values_only=True):
                if not row[1]:
                    continue
                role_id = str(row[1])
                self._roles[role_id] = {
                    "pos_id": int(row[0]) if row[0] else 0,
                    "role_id": role_id,
                    "management": str(row[2]) if row[2] else "",
                    "unit": str(row[3]) if row[3] else "",
                    "role": str(row[4]) if row[4] else "",
                }

            # --- KPI_Indicators ---
            for row in wb["KPI_Indicators"].iter_rows(min_row=2, values_only=True):
                if not row[1]:
                    continue
                role_id = str(row[1])
                formula_type = str(row[7]).strip() if row[7] else "binary_manual"
                if formula_type not in _FORMULA_TO_KPI_TYPE:
                    formula_type = "binary_manual"

                # is_common: Excel хранит "TRUE"/"FALSE" или bool
                raw_common = row[6]
                if isinstance(raw_common, bool):
                    is_common = raw_common
                else:
                    is_common = str(raw_common).upper() == "TRUE"

                # cumulative: аналогично
                raw_cumul = row[10]
                if isinstance(raw_cumul, bool):
                    cumulative = raw_cumul
                else:
                    cumulative = str(raw_cumul).upper() == "TRUE"

                try:
                    weight = int(row[5]) if row[5] is not None else 0
                except (ValueError, TypeError):
                    weight = 0

                item = KpiItem(
                    pos_id=int(row[0]) if row[0] else 0,
                    role_id=role_id,
                    indicator=str(row[2]) if row[2] else "",
                    criterion=str(row[3]) if row[3] else "",
                    plan_value=str(row[4]) if row[4] else "",
                    weight=weight,
                    is_common=is_common,
                    formula_type=formula_type,
                    formula_desc=str(row[8]) if row[8] else "",
                    thresholds=str(row[9]) if row[9] else "",
                    cumulative=cumulative,
                    kpi_type=_FORMULA_TO_KPI_TYPE[formula_type],
                )

                if role_id not in self._indicators:
                    self._indicators[role_id] = []
                self._indicators[role_id].append(item)

            wb.close()
            self._loaded = True
            total = sum(len(v) for v in self._indicators.values())
            logger.info(f"KPI_Mapping загружен: {len(self._roles)} должностей, {total} индикаторов")

        except FileNotFoundError:
            logger.warning(f"KPI_Mapping.xlsx не найден: {KPI_MAPPING_PATH}")
            self._loaded = True
        except Exception as e:
            logger.error(f"Ошибка загрузки KPI_Mapping: {e}")

    def reload(self) -> dict:
        """Сбрасывает кэш и перечитывает KPI_Mapping.xlsx заново."""
        self._loaded = False
        self._roles = {}
        self._indicators = {}
        self._load()
        total = sum(len(v) for v in self._indicators.values())
        return {"roles": len(self._roles), "indicators": total}

    def get_kpi_structure(self, role_id: str) -> KpiStructure:
        """Возвращает KPI-структуру по role_id: три группы + суммарный вес."""
        self._load()
        items = self._indicators.get(role_id, [])

        seen: set[str] = set()
        binary_auto: list[KpiItem] = []
        binary_manual: list[KpiItem] = []
        numeric: list[KpiItem] = []

        for item in items:
            key = item.criterion.strip()
            if key in seen:
                continue
            seen.add(key)
            if item.kpi_type == "binary_auto":
                binary_auto.append(item)
            elif item.kpi_type == "binary_manual":
                binary_manual.append(item)
            else:
                numeric.append(item)

        total_weight = sum(i.weight for i in binary_auto + binary_manual + numeric)

        return KpiStructure(
            role_id=role_id,
            binary_auto=binary_auto,
            binary_manual=binary_manual,
            numeric=numeric,
            total_weight=total_weight,
        )

    def pos_id_to_role_id(self, pos_id: str) -> Optional[str]:
        """Конвертирует числовой pos_id ('4') в role_id ('РУК_ЗАМД_004')."""
        self._load()
        for role_id, info in self._roles.items():
            if str(info.get("pos_id", "")) == str(pos_id):
                return role_id
        return None

    def get_kpi_structure_by_pos_id(self, pos_id: str) -> KpiStructure:
        """Получить структуру KPI по числовому pos_id (хранится в employees.position_id)."""
        role_id = self.pos_id_to_role_id(pos_id)
        if not role_id:
            return KpiStructure(
                role_id="", binary_auto=[], binary_manual=[], numeric=[], total_weight=0
            )
        return self.get_kpi_structure(role_id)

    def get_role_info(self, role_id: str) -> Optional[dict]:
        self._load()
        # Поддержка как role_id ('РУК_ЗАМД_004'), так и pos_id ('4')
        if role_id in self._roles:
            return self._roles[role_id]
        resolved = self.pos_id_to_role_id(role_id)
        return self._roles.get(resolved) if resolved else None

    def get_all_role_ids(self) -> list[str]:
        self._load()
        return list(self._roles.keys())

    def role_exists(self, role_id: str) -> bool:
        self._load()
        return role_id in self._roles

    # --- Обратная совместимость (используются в других местах кода) ---

    def get_kpi_for_role(self, role_id: str) -> list[dict]:
        self._load()
        return [i.model_dump() for i in self._indicators.get(role_id, [])]

    def get_binary_auto_criteria(self, role_id: str) -> list[str]:
        structure = self.get_kpi_structure(role_id)
        seen: set[str] = set()
        result = []
        for item in structure.binary_auto:
            c = item.criterion.strip()
            if c and c not in seen:
                seen.add(c)
                result.append(c)
        return result

    def get_numeric_kpis(self, role_id: str) -> list[dict]:
        structure = self.get_kpi_structure(role_id)
        return [i.model_dump() for i in structure.numeric]

    def get_all_roles(self) -> list[dict]:
        self._load()
        return list(self._roles.values())


kpi_mapping_service = KpiMappingService()
